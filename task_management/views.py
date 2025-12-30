# task_management/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.db.models import Count, Sum, Avg, Max, Min, Q, F
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.utils import timezone

from datetime import date, datetime, timedelta
from decimal import Decimal
import json
import secrets
import string
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.db.models import Max
from .models import TaskType, TaskGroup
# Import models
from .models import (
    TaskGroup, TaskType, TaskPackage, Task, 
    TaskCompletion, PointRequest, Notification,
    ServiceRequest, ClosingReport, Cat, Customer
)
from accounts.models import User
# ============================================
# STEP 1: SEARCH/CREATE CUSTOMER
# ============================================

@login_required
def register_service_step1(request):
    """
    Step 1: Search for existing customer or register new
    Search by: IC, Name, Email, Phone
    """
    if request.method == 'POST':
        search_type = request.POST.get('search_type', 'ic')
        search_query = request.POST.get('search_query', '').strip()
        
        if not search_query:
            messages.error(request, 'Please enter a search term.')
            return redirect('task_management:register_step1')
        
        # Search customers
        customers = None
        if search_type == 'ic':
            customers = Customer.objects.filter(ic_number__icontains=search_query)
        elif search_type == 'name':
            customers = Customer.objects.filter(name__icontains=search_query)
        elif search_type == 'email':
            customers = Customer.objects.filter(email__icontains=search_query)
        elif search_type == 'phone':
            customers = Customer.objects.filter(phone__icontains=search_query)
        
        if customers and customers.exists():
            # Found customers - show selection
            context = {
                'customers': customers,
                'search_query': search_query,
                'search_type': search_type,
            }
            return render(request, 'task_management/registration/results.html', context)
        else:
            # No customer found - show registration form
            context = {
                'search_query': search_query,
                'search_type': search_type,
            }
            return render(request, 'task_management/registration/new_customer.html', context)
    
    # GET request - show search form
    context = {
        'today': date.today(),
    }
    return render(request, 'task_management/registration/search.html', context)


@login_required
def register_service_step1_create(request):
    """
    Step 1B: Create new customer
    """
    if request.method == 'POST':
        # Verify employee ID
        employee_id = request.POST.get('employee_id')
        try:
            staff_user = User.objects.get(employee_id=employee_id, is_active=True)
        except User.DoesNotExist:
            messages.error(request, f'Invalid employee ID: {employee_id}')
            return redirect('task_management:register_step1')
        
        # Get customer data
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        email = request.POST.get('email', '')
        ic_number = request.POST.get('ic_number')
        address = request.POST.get('address', '')
        emergency_contact = request.POST.get('emergency_contact', '')
        
        # Create customer
        try:
            customer = Customer.objects.create(
                name=name,
                phone=phone,
                email=email,
                ic_number=ic_number,
                address=address,
                emergency_contact=emergency_contact,
                registered_by=staff_user,
            )
            
            messages.success(request, f'✓ Customer {customer.customer_id} created successfully!')
            
            # Redirect to step 2 with customer_id
            return redirect('task_management:register_step2', customer_id=customer.customer_id)
        
        except Exception as e:
            messages.error(request, f'Error creating customer: {str(e)}')
            return redirect('task_management:register_step1')
    
    return redirect('task_management:register_step1')


# ============================================
# STEP 2: SELECT/ADD CATS
# ============================================

@login_required
def register_service_step2(request, customer_id):
    """
    Step 2: Select existing cats or add new cat
    Can select multiple cats (will create separate packages)
    """
    customer = get_object_or_404(Customer, customer_id=customer_id)
    cats = customer.cats.filter(is_active=True)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'select_cats':
            # Get selected cat IDs
            selected_cat_ids = request.POST.getlist('selected_cats')
            
            if not selected_cat_ids:
                messages.error(request, 'Please select at least one cat.')
                return redirect('task_management:register_step2', customer_id=customer_id)
            
            # Store in session
            request.session['customer_id'] = customer_id
            request.session['selected_cat_ids'] = selected_cat_ids
            
            # Redirect to step 3
            return redirect('task_management:register_step3')
        
        elif action == 'add_cat':
            # Add new cat
            return redirect('task_management:register_step2_add_cat', customer_id=customer_id)
    
    context = {
        'customer': customer,
        'cats': cats,
    }
    return render(request, 'task_management/registration/select_cats.html', context)


@login_required
def register_service_step2_add_cat(request, customer_id):
    """
    Step 2B: Add new cat for customer
    """
    customer = get_object_or_404(Customer, customer_id=customer_id)
    
    if request.method == 'POST':
        # Verify employee ID
        employee_id = request.POST.get('employee_id')
        try:
            staff_user = User.objects.get(employee_id=employee_id, is_active=True)
        except User.DoesNotExist:
            messages.error(request, f'Invalid employee ID: {employee_id}')
            return redirect('task_management:register_step2_add_cat', customer_id=customer_id)
        
        # Get cat data
        name = request.POST.get('name')
        breed = request.POST.get('breed')
        age = request.POST.get('age')
        gender = request.POST.get('gender')
        color = request.POST.get('color', '')
        vaccination_status = request.POST.get('vaccination_status')
        medical_notes = request.POST.get('medical_notes', '')
        special_requirements = request.POST.get('special_requirements', '')
        
        # Create cat
        try:
            cat = Cat.objects.create(
                name=name,
                owner=customer,
                breed=breed,
                age=age,
                gender=gender,
                color=color,
                vaccination_status=vaccination_status,
                medical_notes=medical_notes,
                special_requirements=special_requirements,
                registered_by=staff_user,
                weight=0,  # Default
            )
            
            messages.success(request, f'✓ Cat {cat.cat_id} - {cat.name} added successfully!')
            
            # Redirect back to step 2
            return redirect('task_management:register_step2', customer_id=customer_id)
        
        except Exception as e:
            messages.error(request, f'Error adding cat: {str(e)}')
            return redirect('task_management:register_step2_add_cat', customer_id=customer_id)
    
    context = {
        'customer': customer,
    }
    return render(request, 'task_management/registration/add_cat.html', context)


# ============================================
# STEP 3: SELECT TASKS
# ============================================

@login_required
def register_service_step3(request):
    """
    Step 3: Select tasks for selected cats
    Same tasks will be applied to all selected cats
    """
    # Get from session
    customer_id = request.session.get('customer_id')
    selected_cat_ids = request.session.get('selected_cat_ids', [])
    
    if not customer_id or not selected_cat_ids:
        messages.error(request, 'Session expired. Please start again.')
        return redirect('task_management:register_step1')
    
    customer = get_object_or_404(Customer, customer_id=customer_id)
    cats = Cat.objects.filter(cat_id__in=selected_cat_ids)
    
    # Get all active task groups and types
    task_groups = TaskGroup.objects.filter(is_active=True).prefetch_related('task_types')
    
    if request.method == 'POST':
        # Verify employee ID
        employee_id = request.POST.get('employee_id')
        try:
            staff_user = User.objects.get(employee_id=employee_id, is_active=True)
        except User.DoesNotExist:
            messages.error(request, f'Invalid employee ID: {employee_id}')
            return redirect('task_management:register_step3')
        
        # Get selected tasks
        selected_task_ids = request.POST.getlist('selected_tasks')
        
        if not selected_task_ids:
            messages.error(request, 'Please select at least one task.')
            return redirect('task_management:register_step3')
        
        # Get task types
        task_types = TaskType.objects.filter(task_type_id__in=selected_task_ids)
        
        # Get package notes
        package_notes = request.POST.get('package_notes', '')
        
        created_packages = []
        
        # Create package for each selected cat
        for cat in cats:
            # Create package
            package = TaskPackage.objects.create(
                cat=cat,
                created_by=staff_user,
                notes=package_notes,
                status='pending',
            )
            
            # Create tasks for this package
            for task_type in task_types:
                # Get date and time for this task
                scheduled_date = request.POST.get(f'date_{task_type.task_type_id}')
                scheduled_time = request.POST.get(f'time_{task_type.task_type_id}', '09:00')
                
                if scheduled_date:
                    Task.objects.create(
                        package=package,
                        task_type=task_type,
                        points=task_type.points,
                        scheduled_date=scheduled_date,
                        scheduled_time=scheduled_time,
                        status='pending',
                    )
            
            # Calculate total points
            package.calculate_total_points()
            
            created_packages.append(package)
            
            # Send email to customer
            try:
                send_registration_email(package, customer)
                package.email_sent = True
                package.email_sent_at = timezone.now()
                package.save()
            except Exception as e:
                print(f"Email error: {e}")
        
        # Create notification for managers
        managers = User.objects.filter(role__in=['manager', 'admin'], is_active=True)
        for manager in managers:
            Notification.objects.create(
                user=manager,
                notification_type='package_created',
                title=f'New Package - {customer.name}',
                message=f'{len(created_packages)} package(s) created by {staff_user.username}. Pending assignment.',
                link='/task-management/unassigned/',
            )
        
        # Clear session
        request.session.pop('customer_id', None)
        request.session.pop('selected_cat_ids', None)
        
        # Success message
        messages.success(
            request,
            f'✓ {len(created_packages)} package(s) created successfully! '
            f'Package IDs: {", ".join([p.package_id for p in created_packages])}. '
            f'Email sent to customer.'
        )
        
        return redirect('task_management:register_step1')
    
    context = {
        'customer': customer,
        'cats': cats,
        'task_groups': task_groups,
        'today': date.today(),
    }
    return render(request, 'task_management/registration/select_tasks.html', context)


# ============================================
# EMAIL FUNCTION
# ============================================

def send_registration_email(package, customer):
    """
    Send registration confirmation email to customer
    """
    tasks = package.tasks.all()
    
    # Prepare context
    context = {
        'customer': customer,
        'package': package,
        'cat': package.cat,
        'tasks': tasks,
        'branch': package.created_by.branch if package.created_by else 'Main Branch',
    }
    
    # Render email template
    html_message = render_to_string('task_management/emails/confirmation.html', context)
    plain_message = f"""
Dear {customer.name},

Thank you for choosing CatzoTeam! Your booking has been confirmed.

Booking ID: {package.package_id}
Cat: {package.cat.name}
Total Tasks: {tasks.count()}
Total Points: {package.total_points}

We will contact you soon with staff assignment details.

Best regards,
CatzoTeam
"""
    
    # Send email
    send_mail(
        subject=f'Booking Confirmed - {package.cat.name}\'s Appointment at CatzoTeam',
        message=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[customer.email],
        html_message=html_message,
        fail_silently=False,
    )


# ============================================
# MANAGER: ASSIGN TASKS
# ============================================
def is_manager(user):
    """Check if user is manager or admin"""
    return user.role in ['manager', 'admin']

@login_required
@user_passes_test(is_manager)
def manager_dashboard(request):
    """Manager dashboard showing pending tasks and assignments"""
    
    # Pending service requests (not yet converted to task packages)
    pending_requests = ServiceRequest.objects.filter(
        status='pending'
    ).select_related('customer', 'requested_by').prefetch_related('cats')
    
    # Pending task packages (created but not assigned)
    pending_packages = TaskPackage.objects.filter(
        status='pending'
    ).select_related('customer', 'created_by').prefetch_related('cats', 'tasks')
    
    # Active task packages (assigned and in progress)
    active_packages = TaskPackage.objects.filter(
        status__in=['assigned', 'in_progress']
    ).select_related('customer', 'assigned_manager').prefetch_related('cats', 'tasks')
    
    # Recent completed tasks (today)
    today = timezone.now().date()
    completed_today = Task.objects.filter(
        completed_at__date=today,
        status='completed'
    ).select_related('task_type', 'assigned_staff', 'cat').order_by('-completed_at')
    
    # Staff availability
    available_staff = User.objects.filter(
        is_active=True,
        role__in=['employee', 'groomer', 'front_desk']
    ).annotate(
        active_tasks=Count('assigned_tasks', filter=Q(assigned_tasks__status__in=['pending', 'in_progress']))
    ).order_by('active_tasks')
    
    # Statistics
    stats = {
        'pending_requests': pending_requests.count(),
        'pending_packages': pending_packages.count(),
        'active_packages': active_packages.count(),
        'completed_today': completed_today.count(),
        'total_points_today': completed_today.aggregate(total=Sum('task_type__points'))['total'] or 0
    }
    
    context = {
        'pending_requests': pending_requests[:10],  # Show latest 10
        'pending_packages': pending_packages[:10],
        'active_packages': active_packages[:10],
        'completed_today': completed_today[:20],
        'available_staff': available_staff[:15],
        'stats': stats
    }
    
    return render(request, 'task_management/manager_dashboard.html', context)


@login_required
@user_passes_test(is_manager)
def assign_task(request, task_id):
    """Manager assigns a task to staff OR self"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard')
    
    task = get_object_or_404(Task, task_id=task_id)
    
    if request.method == 'POST':
        staff_id = request.POST.get('assigned_staff')
        
        # Check if manager is assigning to self
        if staff_id == 'self' or int(staff_id) == request.user.id:
            # Manager self-assignment
            task.assigned_staff = request.user
            task.assigned_by = request.user
            task.assigned_date = timezone.now()
            task.status = 'assigned'
            task.save()
            
            messages.success(
                request,
                f'✅ Task {task.task_id} assigned to yourself! '
                f'Complete it to earn {task.points} points.'
            )
        else:
            # Assign to staff
            try:
                staff = User.objects.get(id=staff_id, is_active=True)
                
                # Branch check for managers (not admin)
                if request.user.role == 'manager':
                    if staff.branch != request.user.branch:
                        messages.error(request, 'You can only assign tasks to staff in your branch.')
                        return redirect('task_management:unassigned_packages')
                
                task.assigned_staff = staff
                task.assigned_by = request.user
                task.assigned_date = timezone.now()
                task.status = 'assigned'
                task.save()
                
                # Update package status
                task.package.update_status()
                
                # Create notification for staff
                Notification.objects.create(
                    user=staff,
                    notification_type='task_assigned',
                    title=f'New Task Assigned',
                    message=f'Task {task.task_id} - {task.task_type.name} for {task.package.cat.name}',
                    link='/task-management/tasks/my-tasks/',
                )
                
                messages.success(request, f'✅ Task {task.task_id} assigned to {staff.username}')
                
            except User.DoesNotExist:
                messages.error(request, 'Invalid staff member.')
        
        return redirect('task_management:unassigned_packages')
    
    return redirect('task_management:unassigned_packages')



@login_required
@user_passes_test(is_manager)
def approve_task_completion(request, task_id):
    """Approve or reject a completed task"""
    
    task = get_object_or_404(Task, id=task_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            task.status = 'completed'
            task.approved_by = request.user
            task.approved_at = timezone.now()
            task.save()
            
            # Award points to staff (we'll implement this in next phase)
            # award_points_to_staff(task)
            
            messages.success(request, f'Task approved! {task.task_type.points} points awarded.')
        
        elif action == 'reject':
            task.status = 'pending'
            task.completed_at = None
            rejection_reason = request.POST.get('rejection_reason')
            task.notes = f'{task.notes}\n\nREJECTED: {rejection_reason}'
            task.save()
            
            messages.warning(request, 'Task rejected and sent back to staff.')
        
        return redirect('task_management:manager_dashboard')
    
    context = {
        'task': task
    }
    
    return render(request, 'task_management/approve_task.html', context)

@login_required
def unassigned_packages(request):
    """Manager views unassigned packages and assigns tasks to staff OR self"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, 'Access denied. Managers only.')
        return redirect('staff_dashboard')
    
    packages = TaskPackage.objects.filter(
        status='pending'
    ).select_related('cat', 'cat__owner', 'created_by').prefetch_related('tasks')
    
    # Get staff based on role
    if request.user.role == 'manager':
        # Manager sees only their branch staff + themselves
        staff = User.objects.filter(
            Q(branch=request.user.branch) | Q(id=request.user.id),
            is_active=True,
            role__in=['staff', 'manager']
        ).order_by('username')
    else:
        # Admin sees all staff + all managers
        staff = User.objects.filter(
            is_active=True,
            role__in=['staff', 'manager']
        ).order_by('branch', 'username')
    
    context = {
        'packages': packages,
        'staff': staff,
        'current_user': request.user,  # So template can show "Assign to Me" option
    }
    
    return render(request, 'task_management/manager/unassigned.html', context)


@login_required
def assign_task(request, task_id):
    """Manager assigns a task to staff"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard')
    
    task = get_object_or_404(Task, task_id=task_id)
    
    if request.method == 'POST':
        staff_id = request.POST.get('assigned_staff')
        
        try:
            staff = User.objects.get(id=staff_id, is_active=True)
            
            task.assigned_staff = staff
            task.assigned_by = request.user
            task.assigned_date = timezone.now()
            task.status = 'assigned'
            task.save()
            
            # Update package status
            task.package.update_status()
            
            # Create notification for staff
            Notification.objects.create(
                user=staff,
                notification_type='task_assigned',
                title=f'New Task Assigned',
                message=f'Task {task.task_id} - {task.task_type.name} for {task.package.cat.name}',
                link='/task-management/my-tasks/',
            )
            
            messages.success(request, f'✓ Task {task.task_id} assigned to {staff.username}')
            
        except User.DoesNotExist:
            messages.error(request, 'Invalid staff member.')
        
        return redirect('task_management:unassigned_packages')
    
    return redirect('task_management:unassigned_packages')


# ============================================
# STAFF: MY TASKS
# ============================================

@login_required
def my_tasks(request):
    """Staff views their assigned tasks"""
    today = date.today()
    
    # Today's tasks
    today_tasks = Task.objects.filter(
        assigned_staff=request.user,
        scheduled_date=today,
        status__in=['assigned', 'in_progress']
    ).select_related('package__cat', 'package__cat__owner', 'task_type')
    
    # Upcoming tasks
    upcoming_tasks = Task.objects.filter(
        assigned_staff=request.user,
        scheduled_date__gt=today,
        status__in=['assigned', 'in_progress']
    ).select_related('package__cat', 'package__cat__owner', 'task_type').order_by('scheduled_date')[:10]
    
    # Completed tasks today
    completed_today = Task.objects.filter(
        assigned_staff=request.user,
        completed_at__date=today,
        status='completed'
    ).select_related('package__cat', 'task_type')
    
    # Calculate today's points
    today_points = sum(t.points for t in completed_today)
    
    context = {
        'today_tasks': today_tasks,
        'upcoming_tasks': upcoming_tasks,
        'completed_today': completed_today,
        'today_points': today_points,
        'today': today,
    }
    
    return render(request, 'task_management/staff/my_tasks.html', context)


@login_required
def complete_task(request, task_id):
    """Staff marks task as completed and earns points"""
    task = get_object_or_404(Task, task_id=task_id)
    
    if task.assigned_staff != request.user:
        messages.error(request, 'This task is not assigned to you.')
        return redirect('task_management:my_tasks')
    
    if request.method == 'POST':
        completion_notes = request.POST.get('completion_notes', '')
        
        # Mark task as completed
        task.status = 'completed'
        task.completed_at = timezone.now()
        task.save()
        
        # Create completion record
        completion = TaskCompletion.objects.create(
            task=task,
            completed_by=request.user,
            completion_notes=completion_notes,
        )
        
        # Award points immediately
        completion.award_points()
        
        # Update package status
        task.package.update_status()
        
        messages.success(
            request,
            f'✓ Task completed! You earned {task.points} points.'
        )
        
        return redirect('task_management:my_tasks')
    
    return redirect('task_management:my_tasks')

# Add this to your task_management/views.py

@login_required
def my_tasks(request):
    """Staff view - Show all assigned tasks"""
    user = request.user
    
    # Get tasks by status
    assigned_tasks = Task.objects.filter(
        assigned_staff=user,
        status='assigned'
    ).select_related(
        'task_type',
        'package__cat',
        'package__cat__owner'
    ).order_by('-id')[:20]
    
    in_progress_tasks = Task.objects.filter(
        assigned_staff=user,
        status='in_progress'
    ).select_related(
        'task_type',
        'package__cat',
        'package__cat__owner'
    ).order_by('-id')[:20]
    
    pending_approval = Task.objects.filter(
        assigned_staff=user,
        status='submitted'
    ).select_related(
        'task_type',
        'package__cat',
        'package__cat__owner'
    ).order_by('-id')[:20]
    
    completed_tasks = Task.objects.filter(
        assigned_staff=user,
        status='completed'
    ).select_related(
        'task_type',
        'package__cat',
        'package__cat__owner'
    ).order_by('-id')[:10]
    
    # Count by status
    tasks_count = {
        'assigned': assigned_tasks.count(),
        'in_progress': in_progress_tasks.count(),
        'submitted': pending_approval.count(),
        'completed': Task.objects.filter(assigned_staff=user, status='completed').count(),
    }
    
    context = {
        'assigned_tasks': assigned_tasks,
        'in_progress_tasks': in_progress_tasks,
        'pending_approval': pending_approval,
        'completed_tasks': completed_tasks,
        'tasks_count': tasks_count,
    }
    
    return render(request, 'task_management/staff/my_tasks.html', context)


@login_required
def complete_task(request, task_id):
    """Staff - Complete a task"""
    task = get_object_or_404(Task, task_id=task_id, assigned_staff=request.user)
    
    if request.method == 'POST':
        # Update task status
        task.status = 'submitted'  # Or 'completed' if no approval needed
        task.completed_at = timezone.now()
        
        # Add completion notes if provided
        notes = request.POST.get('completion_notes', '')
        if notes:
            task.notes = f"{task.notes}\n\n[COMPLETED]: {notes}" if task.notes else f"[COMPLETED]: {notes}"
        
        task.save()
        
        messages.success(request, f'✅ Task "{task.task_type.name}" submitted for approval!')
        return redirect('task_management:my_tasks')
    
    context = {'task': task}
    return render(request, 'task_management/staff/complete_task.html', context)

@login_required
def my_tasks_view(request):
    """Staff/Manager view their own assigned tasks with filters"""
    user = request.user
    
    # Get tasks by status - FIXED: Use package__cat instead of cat
    assigned_tasks = Task.objects.filter(
        assigned_staff=user,
        status='assigned'
    ).select_related(
        'task_type',
        'package__cat',           # ← FIXED: was 'cat'
        'package__cat__owner'     # ← FIXED: was 'cat__owner'
    ).order_by('-id')[:20]
    
    in_progress_tasks = Task.objects.filter(
        assigned_staff=user,
        status='in_progress'
    ).select_related(
        'task_type',
        'package__cat',           # ← FIXED
        'package__cat__owner'     # ← FIXED
    ).order_by('-id')[:20]
    
    # For staff: pending approval. For manager: none (self-approved)
    if user.role == 'staff':
        pending_approval = Task.objects.filter(
            assigned_staff=user,
            status='submitted'
        ).select_related(
            'task_type', 
            'package__cat',       # ← FIXED
            'package__cat__owner' # ← FIXED
        ).order_by('-id')[:20]
    else:
        pending_approval = []
    
    completed_tasks = Task.objects.filter(
        assigned_staff=user,
        status='completed'
    ).select_related(
        'task_type', 
        'package__cat',           # ← FIXED
        'package__cat__owner'     # ← FIXED
    ).order_by('-completed_at')[:10]
    
    # Count by status
    tasks_count = {
        'assigned': assigned_tasks.count(),
        'in_progress': in_progress_tasks.count(),
        'submitted': len(pending_approval) if user.role == 'staff' else 0,
        'completed': Task.objects.filter(assigned_staff=user, status='completed').count(),
        'total': Task.objects.filter(assigned_staff=user).count(),
    }
    
    context = {
        'assigned_tasks': assigned_tasks,
        'in_progress_tasks': in_progress_tasks,
        'pending_approval': pending_approval,
        'completed_tasks': completed_tasks,
        'tasks_count': tasks_count,
    }
    
    return render(request, 'task_management/staff/my_tasks_new.html', context)

@login_required
def complete_task_with_proof(request, task_id):
    """Complete a task with notes and photo proof - AUTOMATED"""
    task = get_object_or_404(Task, task_id=task_id, assigned_staff=request.user)
    
    if request.method == 'POST':
        # Get completion notes
        notes = request.POST.get('completion_notes', '').strip()
        
        # Handle image uploads
        uploaded_files = request.FILES.getlist('task_images')
        image_paths = []
        
        for uploaded_file in uploaded_files:
            # Save image to media folder
            file_name = f'task_proofs/{timezone.now().strftime("%Y/%m/%d")}/{task.task_id}_{uploaded_file.name}'
            path = default_storage.save(file_name, ContentFile(uploaded_file.read()))
            image_paths.append(path)
        
        # AUTOMATION: Auto-complete for regular tasks (no approval needed)
        task.status = 'completed'
        task.completed_at = timezone.now()
        task.approved_by = request.user  # Self-approved
        task.approved_at = timezone.now()
        
        # Add notes to task
        if notes:
            task.notes = f"{task.notes}\n\n[COMPLETED {timezone.now().strftime('%Y-%m-%d %H:%M')}]: {notes}" if task.notes else f"[COMPLETED]: {notes}"
        
        task.save()
        
        # Create completion record
        completion = TaskCompletion.objects.create(
            task=task,
            completed_by=request.user,
            completion_notes=notes,
            photo_proof=','.join(image_paths) if image_paths else '',
        )
        
        # Award points immediately
        completion.award_points()
        
        # Update package status
        task.package.update_status()
        
        # Notify manager
        if task.assigned_by and task.assigned_by != request.user:
            Notification.objects.create(
                user=task.assigned_by,
                notification_type='task_completed',
                title=f'Task Completed',
                message=f'{request.user.username} completed task {task.task_id} - {task.task_type.name}',
                link='/task-management/manager/staff-tasks/',
            )
        
        messages.success(
            request,
            f'🎉 Task completed! You earned {task.points} points instantly!'
        )
        
        return redirect('task_management:my_tasks_new')
    
    context = {'task': task}
    return render(request, 'task_management/staff/complete_task_new.html', context)


@login_required
def request_custom_points(request):
    """Request points for tasks not in predefined list - FIXED"""
    if request.method == 'POST':
        task_name = request.POST.get('task_name')
        task_description = request.POST.get('task_description')
        requested_points = request.POST.get('requested_points')
        task_date_str = request.POST.get('task_date')
        
        # Parse date
        task_date = datetime.strptime(task_date_str, '%Y-%m-%d').date()
        
        # Handle file uploads
        uploaded_files = request.FILES.getlist('proof_images')
        proof_paths = []
        
        for uploaded_file in uploaded_files:
            file_name = f'point_requests/{timezone.now().strftime("%Y/%m/%d")}/{request.user.username}_{uploaded_file.name}'
            path = default_storage.save(file_name, ContentFile(uploaded_file.read()))
            proof_paths.append(path)
        
        # Create point request using PointRequest model
        point_request = PointRequest.objects.create(
            staff=request.user,
            task_type=None,  # Custom task, no predefined type
            points_requested=Decimal(requested_points),
            date_completed=task_date,
            reason='not_in_system',
            reason_details=f"Task: {task_name}\n\nDescription: {task_description}",
            proof_photo=None,  # We'll handle this separately if needed
            approval_status='pending'
        )
        
        messages.success(
            request,
            f'✅ Point request submitted! Request ID: {point_request.request_id}. '
            f'Waiting for manager/admin approval.'
        )
        return redirect('task_management:my_point_requests')
    
    # GET - Show form
    task_types = TaskType.objects.filter(is_active=True).order_by('group__name', 'name')
    
    context = {
        'task_types': task_types,
        'today': date.today(),
    }
    return render(request, 'task_management/staff/request_points.html', context)


@login_required
def my_point_requests(request):
    """View my custom point requests - FIXED"""
    requests_list = PointRequest.objects.filter(
        staff=request.user
    ).select_related('task_type', 'approved_by').order_by('-created_at')
    
    counts = {
        'pending': requests_list.filter(approval_status='pending').count(),
        'approved': requests_list.filter(approval_status='approved').count(),
        'rejected': requests_list.filter(approval_status='rejected').count(),
        'total': requests_list.count(),
    }
    
    context = {
        'requests': requests_list,
        'counts': counts,
    }
    return render(request, 'task_management/staff/my_point_requests.html', context)


# ==================== MANAGER VIEWS (BRANCH TASKS) ====================

@login_required
def staff_tasks_branch(request):
    """Manager views staff tasks from their branch only"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, 'Access denied.')
        return redirect('task_management:my_tasks_new')
    
    # Get staff from manager's branch
    if request.user.role == 'manager':
        manager_branch = request.user.branch
        branch_staff = User.objects.filter(
            role='staff',
            branch=manager_branch,
            is_active=True
        )
    else:
        # Admin sees all
        branch_staff = User.objects.filter(
            role='staff',
            is_active=True
        )
        manager_branch = 'All Branches'
    
    # Filter parameters
    status_filter = request.GET.get('status', 'all')
    staff_filter = request.GET.get('staff', 'all')
    
    # Base query - FIXED: Use package__cat instead of cat
    tasks = Task.objects.filter(
        assigned_staff__in=branch_staff
    ).select_related(
        'task_type',
        'assigned_staff',
        'package__cat',           # ← FIXED!
        'package__cat__owner'     # ← FIXED!
    )
    
    # Apply filters
    if status_filter != 'all':
        tasks = tasks.filter(status=status_filter)
    
    if staff_filter != 'all':
        tasks = tasks.filter(assigned_staff_id=staff_filter)
    
    tasks = tasks.order_by('-id')[:50]
    
    # Statistics
    stats = {
        'assigned': Task.objects.filter(assigned_staff__in=branch_staff, status='assigned').count(),
        'in_progress': Task.objects.filter(assigned_staff__in=branch_staff, status='in_progress').count(),
        'submitted': Task.objects.filter(assigned_staff__in=branch_staff, status='submitted').count(),
        'completed': Task.objects.filter(assigned_staff__in=branch_staff, status='completed').count(),
        'total': Task.objects.filter(assigned_staff__in=branch_staff).count(),
    }
    
    context = {
        'tasks': tasks,
        'branch_staff': branch_staff,
        'stats': stats,
        'status_filter': status_filter,
        'staff_filter': staff_filter,
        'manager_branch': manager_branch,
    }
    
    return render(request, 'task_management/manager/staff_tasks_branch.html', context)


# ==================== ADMIN VIEWS ====================

@login_required
def admin_all_tasks_monitor(request):
    """Admin monitors ALL tasks from all branches"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin role required.')
        return redirect('dashboard:admin_dashboard')
    
    # Filter parameters
    branch_filter = request.GET.get('branch', 'all')
    status_filter = request.GET.get('status', 'all')
    staff_filter = request.GET.get('staff', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base query - FIXED: Use package__cat instead of cat
    tasks = Task.objects.select_related(
        'task_type',
        'assigned_staff',
        'package__cat',          # ← FIXED!
        'package__cat__owner',   # ← FIXED!
        'approved_by'
    )
    
    # Apply filters
    if branch_filter != 'all':
        tasks = tasks.filter(assigned_staff__branch=branch_filter)
    
    if status_filter != 'all':
        tasks = tasks.filter(status=status_filter)
    
    if staff_filter != 'all':
        tasks = tasks.filter(assigned_staff_id=staff_filter)
    
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        tasks = tasks.filter(assigned_date__gte=date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        tasks = tasks.filter(assigned_date__lte=date_to_obj)
    
    tasks = tasks.order_by('-id')[:100]
    
    # Get all staff for filter
    all_staff = User.objects.filter(
        Q(role='staff') | Q(role='manager'),
        is_active=True
    ).order_by('username')
    
    # Statistics
    stats = {
        'pending': Task.objects.filter(status='pending').count(),
        'assigned': Task.objects.filter(status='assigned').count(),
        'in_progress': Task.objects.filter(status='in_progress').count(),
        'submitted': Task.objects.filter(status='submitted').count(),
        'completed': Task.objects.filter(status='completed').count(),
        'total': Task.objects.count(),
    }
    
    # Branch statistics
    branch_stats = {}
    for branch_code, branch_name in User.BRANCH_CHOICES:
        branch_stats[branch_code] = {
            'name': branch_name,
            'total': Task.objects.filter(assigned_staff__branch=branch_code).count(),
            'completed': Task.objects.filter(assigned_staff__branch=branch_code, status='completed').count(),
        }
    
    context = {
        'tasks': tasks,
        'all_staff': all_staff,
        'stats': stats,
        'branch_stats': branch_stats,
        'filters': {
            'branch': branch_filter,
            'status': status_filter,
            'staff': staff_filter,
            'date_from': date_from,
            'date_to': date_to,
        },
        'branches': User.BRANCH_CHOICES,
    }
    
    return render(request, 'task_management/admin/all_tasks_admin.html', context)


@login_required
def admin_point_requests(request):
    """Admin views and approves/rejects custom point requests"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('dashboard:admin_dashboard')
    
    # Filter by status
    status_filter = request.GET.get('status', 'pending')
    
    requests_list = PointRequest.objects.select_related('staff', 'approved_by')
    
    if status_filter != 'all':
        requests_list = requests_list.filter(approval_status=status_filter)
    
    requests_list = requests_list.order_by('-created_at')[:50]
    
    # Statistics
    stats = {
        'pending': PointRequest.objects.filter(approval_status='pending').count(),
        'approved': PointRequest.objects.filter(approval_status='approved').count(),
        'rejected': PointRequest.objects.filter(approval_status='rejected').count(),
        'total': PointRequest.objects.count(),
    }
    
    context = {
        'requests': requests_list,
        'stats': stats,
        'status_filter': status_filter,
    }
    
    return render(request, 'task_management/admin/point_requests_admin.html', context)


@login_required
def review_point_request(request, pk):  # Changed from request_id to pk
    """Admin/Manager reviews and approves/rejects point request"""
    if request.user.role not in ['admin', 'manager']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:staff_dashboard')
    
    point_request = get_object_or_404(PointRequest, pk=pk)  # Changed to pk
    
    # Manager can only review requests from their branch
    if request.user.role == 'manager':
        if point_request.staff.branch != request.user.branch:
            messages.error(request, 'You can only review requests from your branch staff.')
            return redirect('task_management:admin_point_requests')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('review_notes', '')
        
        if action == 'approve':
            awarded_points = request.POST.get('awarded_points', point_request.points_requested)
            
            # Use the model's approve method
            point_request.approve(
                admin_user=request.user,
                notes=notes,
                awarded_points=Decimal(str(awarded_points))
            )
            
            messages.success(
                request,
                f'✅ Request approved! {point_request.staff.username} will receive '
                f'{awarded_points} points on {point_request.date_completed}.'
            )
        
        elif action == 'reject':
            # Use the model's reject method
            point_request.reject(
                admin_user=request.user,
                notes=notes
            )
            
            messages.warning(request, 'Request rejected.')
        
        return redirect('task_management:admin_point_requests')
    
    context = {'point_request': point_request}
    return render(request, 'task_management/admin/review_point_request.html', context)

@login_required
def submit_closing_report(request):
    """Manager submits daily closing report with ONE image"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, 'Access denied. Manager role required.')
        return redirect('dashboard:staff_dashboard')
    
    if request.method == 'POST':
        # Get form data
        report_date_str = request.POST.get('date')
        grooming_count = request.POST.get('grooming_count', 0)
        boarding_count = request.POST.get('boarding_count', 0)
        total_customers = request.POST.get('total_customers', 0)
        payment_record = request.POST.get('payment_record_amount', 0)
        payment_receipt = request.POST.get('payment_receipt_amount', 0)
        compliance_system = request.POST.get('compliance_all_paid_through_system') == 'yes'
        compliance_free = request.POST.get('compliance_free_services_today') == 'yes'
        notes = request.POST.get('notes', '')
        
        report_date_obj = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        
        existing_report = ClosingReport.objects.filter(
            date=report_date_obj,
            branch=request.user.branch
        ).first()
        
        if existing_report:
            messages.error(request, f'Closing report for {report_date_str} already submitted!')
            return redirect('task_management:my_closing_reports')
        
        # ✅ CHANGE THIS LINE - use payment_proof_photo instead
        payment_proof_photo = request.FILES.get('payment_proof_photo')
        
        if not payment_proof_photo:
            messages.error(request, 'Payment proof image is required!')
            return redirect('task_management:submit_closing_report')
        
        # Create closing report
        report = ClosingReport.objects.create(
            date=report_date_obj,
            branch=request.user.branch,
            submitted_by=request.user,
            grooming_count=int(grooming_count),
            boarding_count=int(boarding_count),
            total_customers=int(total_customers),
            payment_record_amount=Decimal(payment_record),
            payment_receipt_amount=Decimal(payment_receipt),
            payment_proof_photo=payment_proof_photo,  # ✅ CHANGED
            compliance_all_paid_through_system=compliance_system,
            compliance_free_services_today=compliance_free,
            notes=notes
        )
        
        messages.success(request, f'✅ Closing report {report.report_id} submitted successfully!')
        return redirect('task_management:my_closing_reports')
    
    context = {
        'today': date.today(),
        'branch': request.user.get_branch_display(),
    }
    return render(request, 'task_management/closing_reports/submit_report.html', context)

# ============================================
# MANAGER: VIEW MY CLOSING REPORTS
# ============================================

@login_required
def my_closing_reports(request):
    """Manager views their branch's closing reports"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:staff_dashboard')
    
    # Get reports for manager's branch
    reports = ClosingReport.objects.filter(
        branch=request.user.branch
    ).order_by('-date')[:30]  # Last 30 reports
    
    # Calculate statistics
    last_30_days = date.today() - timedelta(days=30)
    stats = ClosingReport.objects.filter(
        branch=request.user.branch,
        date__gte=last_30_days
    ).aggregate(
        total_revenue=Sum('revenue_total'),
        total_customers=Sum('total_customers'),
        total_grooming=Sum('grooming_count'),
        total_boarding=Sum('boarding_count'),
        avg_transaction=Avg('revenue_total'),
        report_count=Count('id')
    )
    
    context = {
        'reports': reports,
        'stats': stats,
        'branch': request.user.get_branch_display(),
    }
    return render(request, 'task_management/closing_reports/my_reports.html', context)


# ============================================
# MANAGER: ANALYTICS DASHBOARD
# ============================================

@login_required
def manager_analytics_dashboard(request):
    """Manager views sales analytics and projections for their branch"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:staff_dashboard')
    
    # Get time period from query param (default: 30 days)
    period = request.GET.get('period', '30')
    period_days = int(period)
    
    start_date = date.today() - timedelta(days=period_days)
    
    # Get closing reports for the period
    reports = ClosingReport.objects.filter(
        branch=request.user.branch,
        date__gte=start_date
    ).order_by('date')
    
    # Prepare data for charts
    dates = []
    revenues = []
    customers = []
    grooming_counts = []
    boarding_counts = []
    
    for report in reports:
        dates.append(report.date.strftime('%Y-%m-%d'))
        revenues.append(float(report.revenue_total))
        customers.append(report.total_customers)
        grooming_counts.append(report.grooming_count)
        boarding_counts.append(report.boarding_count)
    
    # Calculate 30-day projection
    last_30_days = date.today() - timedelta(days=30)
    last_30_reports = ClosingReport.objects.filter(
        branch=request.user.branch,
        date__gte=last_30_days
    )
    
    avg_stats = last_30_reports.aggregate(
        avg_revenue=Avg('revenue_total'),
        avg_customers=Avg('total_customers'),
        avg_grooming=Avg('grooming_count'),
        avg_boarding=Avg('boarding_count')
    )
    
    # Generate projection for next 30 days
    projection_dates = []
    projection_revenues = []
    
    for i in range(1, 31):
        future_date = date.today() + timedelta(days=i)
        projection_dates.append(future_date.strftime('%Y-%m-%d'))
        projection_revenues.append(float(avg_stats['avg_revenue'] or 0))
    
    # Calculate summary stats
    summary = last_30_reports.aggregate(
        total_revenue=Sum('revenue_total'),
        total_customers=Sum('total_customers'),
        total_reports=Count('id')
    )
    
    context = {
        'branch': request.user.get_branch_display(),
        'period': period,
        'dates': json.dumps(dates),
        'revenues': json.dumps(revenues),
        'customers': json.dumps(customers),
        'grooming_counts': json.dumps(grooming_counts),
        'boarding_counts': json.dumps(boarding_counts),
        'projection_dates': json.dumps(projection_dates),
        'projection_revenues': json.dumps(projection_revenues),
        'avg_stats': avg_stats,
        'summary': summary,
    }
    return render(request, 'task_management/closing_reports/manager_analytics.html', context)


# ============================================
# ADMIN: VIEW ALL CLOSING REPORTS
# ============================================

@login_required
def admin_all_closing_reports(request):
    """Admin views closing reports from all branches"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin role required.')
        return redirect('dashboard:admin_dashboard')
    
    # Filters
    branch_filter = request.GET.get('branch', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base query
    reports = ClosingReport.objects.all()
    
    # Apply filters
    if branch_filter != 'all':
        reports = reports.filter(branch=branch_filter)
    
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        reports = reports.filter(date__gte=date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        reports = reports.filter(date__lte=date_to_obj)
    
    reports = reports.order_by('-date')[:50]
    
    # Get all branches for filter
    branches = User.BRANCH_CHOICES
    
    # Calculate totals
    last_30_days = date.today() - timedelta(days=30)
    totals = ClosingReport.objects.filter(
        date__gte=last_30_days
    ).aggregate(
        total_revenue=Sum('revenue_total'),
        total_customers=Sum('total_customers'),
        total_reports=Count('id')
    )
    
    context = {
        'reports': reports,
        'branches': branches,
        'branch_filter': branch_filter,
        'date_from': date_from,
        'date_to': date_to,
        'totals': totals,
    }
    return render(request, 'task_management/closing_reports/admin_all_reports.html', context)


# ============================================
# ADMIN: ANALYTICS DASHBOARD (ALL BRANCHES)
# ============================================

def admin_analytics_dashboard(request):
    """Admin analytics for all branches - FIXED to show all branches even with no data"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('dashboard:admin_dashboard')
    
    period = request.GET.get('period', '30')
    period_days = int(period)
    
    start_date = date.today() - timedelta(days=period_days)
    end_date = date.today()
    
    # Generate complete date range for the period
    date_range = []
    current_date = start_date
    while current_date <= end_date:
        date_range.append(current_date)
        current_date += timedelta(days=1)
    
    # Format dates for JSON
    date_labels = [d.strftime('%Y-%m-%d') for d in date_range]
    
    # Build data for each branch
    branches_data = {}
    
    for branch_code, branch_name in User.BRANCH_CHOICES:
        # Get all reports for this branch in the period
        reports = ClosingReport.objects.filter(
            branch=branch_code,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date')
        
        # Create a lookup dictionary: date -> report
        reports_by_date = {
            report.date: report for report in reports
        }
        
        # Build arrays with 0 for missing dates
        revenues = []
        customers = []
        
        for current_date in date_range:
            if current_date in reports_by_date:
                report = reports_by_date[current_date]
                revenues.append(float(report.revenue_total))
                customers.append(report.total_customers)
            else:
                # No report for this date - use 0
                revenues.append(0)
                customers.append(0)
        
        # Calculate 30-day average (only from actual reports, not zeros)
        last_30_days = date.today() - timedelta(days=30)
        recent_reports = ClosingReport.objects.filter(
            branch=branch_code,
            date__gte=last_30_days
        )
        
        if recent_reports.exists():
            avg_revenue = recent_reports.aggregate(avg=Avg('revenue_total'))['avg'] or 0
        else:
            avg_revenue = 0
        
        branches_data[branch_code] = {
            'name': branch_name,
            'dates': date_labels,  # Same for all branches
            'revenues': revenues,
            'customers': customers,
            'avg_revenue': float(avg_revenue),
        }
    
    # Generate projection dates (next 30 days)
    projection_dates = []
    for i in range(1, 31):
        future_date = date.today() + timedelta(days=i)
        projection_dates.append(future_date.strftime('%Y-%m-%d'))
    
    # Calculate system-wide stats
    last_30_days = date.today() - timedelta(days=30)
    system_stats = ClosingReport.objects.filter(
        date__gte=last_30_days
    ).aggregate(
        total_revenue=Sum('revenue_total'),
        total_customers=Sum('total_customers'),
        total_reports=Count('id'),
        avg_revenue=Avg('revenue_total')
    )
    
    # Ensure defaults if no data
    if system_stats['total_revenue'] is None:
        system_stats['total_revenue'] = 0
    if system_stats['total_customers'] is None:
        system_stats['total_customers'] = 0
    if system_stats['total_reports'] is None:
        system_stats['total_reports'] = 0
    if system_stats['avg_revenue'] is None:
        system_stats['avg_revenue'] = 0
    
    context = {
        'period': period,
        'branches_data': json.dumps(branches_data),
        'projection_dates': json.dumps(projection_dates),
        'system_stats': system_stats,
    }
    return render(request, 'task_management/closing_reports/admin_analytics.html', context)

# ============================================
# ADMIN: EXPORT TO EXCEL
# ============================================

@login_required
def export_reports_excel(request):
    """Export closing reports to Excel"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('dashboard:admin_dashboard')
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get filters
    branch_filter = request.GET.get('branch', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Query reports
    reports = ClosingReport.objects.all().order_by('-date')
    
    if branch_filter != 'all':
        reports = reports.filter(branch=branch_filter)
    if date_from:
        reports = reports.filter(date__gte=date_from)
    if date_to:
        reports = reports.filter(date__lte=date_to)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Closing Reports'
    
    # Header style
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Headers
    headers = [
        'Report ID', 'Date', 'Branch', 'Submitted By',
        'Grooming Count', 'Boarding Count', 'Total Customers',
        'Payment Record (RM)', 'Payment Receipt (RM)', 'Revenue Total (RM)',
        'All Paid Through System', 'Free Services Today', 'Notes'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_num, report in enumerate(reports, 2):
        ws.cell(row=row_num, column=1, value=report.report_id)
        ws.cell(row=row_num, column=2, value=report.date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=report.branch)
        ws.cell(row=row_num, column=4, value=report.submitted_by.username if report.submitted_by else '')
        ws.cell(row=row_num, column=5, value=report.grooming_count)
        ws.cell(row=row_num, column=6, value=report.boarding_count)
        ws.cell(row=row_num, column=7, value=report.total_customers)
        ws.cell(row=row_num, column=8, value=float(report.payment_record_amount))
        ws.cell(row=row_num, column=9, value=float(report.payment_receipt_amount))
        ws.cell(row=row_num, column=10, value=float(report.revenue_total))
        ws.cell(row=row_num, column=11, value='Yes' if report.compliance_all_paid_through_system else 'No')
        ws.cell(row=row_num, column=12, value='Yes' if report.compliance_free_services_today else 'No')
        ws.cell(row=row_num, column=13, value=report.notes or '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=closing_reports_{date.today()}.xlsx'
    
    wb.save(response)
    return response


# ============================================
# MANAGE TASKS - TASK TYPES & GROUPS
# ============================================

@login_required
def admin_manage_tasks_page(request):
    """Admin manages task types and groups"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin only.')
        return redirect('dashboard:admin_dashboard')
    
    # Get all groups with task type counts
    groups = TaskGroup.objects.annotate(
        task_count=Count('task_types')
    ).order_by('order', 'name')
    
    # Get all task types with usage count
    task_types = TaskType.objects.annotate(
        usage_count=Count('tasks')
    ).select_related('group').order_by('group__order', 'order', 'name')
    
    # Get selected group from query param
    selected_group = request.GET.get('group', 'all')
    
    # Filter types by group if selected
    if selected_group != 'all':
        task_types = task_types.filter(group_id=selected_group)
    
    # Check if "Uncategorized" group exists, create if not
    uncategorized_group, created = TaskGroup.objects.get_or_create(
        name='Uncategorized',
        defaults={
            'description': 'Uncategorized tasks',
            'is_active': True,
            'order': 9999
        }
    )
    
    context = {
        'groups': groups,
        'task_types': task_types,
        'selected_group': selected_group,
        'uncategorized_group': uncategorized_group,
    }
    return render(request, 'dashboard/admin/manage_tasks.html', context)


@login_required
@require_http_methods(["POST"])
def ajax_create_task_group(request):
    """AJAX: Create new task group"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)
        
        # Check if name exists
        if TaskGroup.objects.filter(name=name).exists():
            return JsonResponse({'success': False, 'error': 'Group name already exists'}, status=400)
        
        # Get max order + 1
        max_order = TaskGroup.objects.aggregate(Max('order'))['order__max'] or 0
        
        group = TaskGroup.objects.create(
            name=name,
            description=description,
            is_active=True,
            order=max_order + 1
        )
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='create',
            model_type='taskgroup',
            object_id=group.group_id,
            object_repr=f"TaskGroup: {group.name}",
            changes={'name': name, 'description': description},
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'group': {
                'id': group.group_id,
                'name': group.name,
                'description': group.description,
                'task_count': 0
            }
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_update_task_group(request, group_id):
    """AJAX: Update task group"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        group = get_object_or_404(TaskGroup, group_id=group_id)
        data = json.loads(request.body)
        
        old_values = {
            'name': group.name,
            'description': group.description,
            'is_active': group.is_active
        }
        
        # Update fields
        if 'name' in data:
            name = data['name'].strip()
            if not name:
                return JsonResponse({'success': False, 'error': 'Name cannot be empty'}, status=400)
            group.name = name
        
        if 'description' in data:
            group.description = data['description'].strip()
        
        if 'is_active' in data:
            group.is_active = data['is_active']
        
        group.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='update',
            model_type='taskgroup',
            object_id=group.group_id,
            object_repr=f"TaskGroup: {group.name}",
            changes={'old': old_values, 'new': {
                'name': group.name,
                'description': group.description,
                'is_active': group.is_active
            }},
            request=request
        )
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_delete_task_group(request, group_id):
    """AJAX: Delete task group (moves types to Uncategorized)"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        group = get_object_or_404(TaskGroup, group_id=group_id)
        
        # Get or create Uncategorized group
        uncategorized_group, created = TaskGroup.objects.get_or_create(
            name='Uncategorized',
            defaults={
                'description': 'Uncategorized tasks',
                'is_active': True,
                'order': 9999
            }
        )
        
        # Move all task types to Uncategorized
        moved_count = group.task_types.update(group=uncategorized_group)
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='delete',
            model_type='taskgroup',
            object_id=group.group_id,
            object_repr=f"TaskGroup: {group.name}",
            notes=f"Moved {moved_count} task types to Uncategorized",
            request=request
        )
        
        group.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Group deleted. {moved_count} task types moved to Uncategorized.'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_reorder_group(request, group_id):
    """AJAX: Move group up or down"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        group = get_object_or_404(TaskGroup, group_id=group_id)
        data = json.loads(request.body)
        direction = data.get('direction')  # 'up' or 'down'
        
        if direction == 'up':
            # Swap with previous
            previous = TaskGroup.objects.filter(order__lt=group.order).order_by('-order').first()
            if previous:
                group.order, previous.order = previous.order, group.order
                group.save()
                previous.save()
        
        elif direction == 'down':
            # Swap with next
            next_group = TaskGroup.objects.filter(order__gt=group.order).order_by('order').first()
            if next_group:
                group.order, next_group.order = next_group.order, group.order
                group.save()
                next_group.save()
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# TASK TYPE AJAX OPERATIONS
# ============================================

@login_required
@require_http_methods(["POST"])
def ajax_create_task_type(request):
    """AJAX: Create new task type"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'})
    
    try:
        data = json.loads(request.body)
        print(f"DEBUG: Received data: {data}")  # Debug log
        
        # Get the group by group_id (not id)
        group_id = data.get('group')
        print(f"DEBUG: Looking for group: {group_id}")  # Debug log
        
        try:
            group = TaskGroup.objects.get(group_id=group_id)
            print(f"DEBUG: Found group: {group}")  # Debug log
        except TaskGroup.DoesNotExist:
            return JsonResponse({'success': False, 'error': f'Group with ID {group_id} not found'})
        
        # Get max order for this group
        max_order = TaskType.objects.filter(group=group).aggregate(
            max_order=Max('order')
        )['max_order'] or 0
        
        print(f"DEBUG: Creating TaskType with name={data.get('name')}, points={data.get('points')}")  # Debug log
        
        # Create task type
        task_type = TaskType.objects.create(
            name=data.get('name'),
            group=group,  # Pass the group object, not the ID
            points=int(data.get('points', 0)),
            price=float(data.get('price')) if data.get('price') else None,
            description=data.get('description', ''),
            category=data.get('category', ''),
            is_active=data.get('is_active', True),
            order=max_order + 10
        )
        
        print(f"DEBUG: Successfully created task type: {task_type.task_type_id}")  # Debug log
        
        return JsonResponse({
            'success': True,
            'message': 'Task type created successfully',
            'task_type_id': task_type.task_type_id
        })
        
    except json.JSONDecodeError as e:
        print(f"ERROR: JSON decode error: {e}")  # Debug log
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"ERROR: Exception creating task type:")  # Debug log
        print(error_trace)  # Debug log
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_http_methods(["POST"])
def ajax_update_task_type(request, type_id):
    """AJAX: Update task type (inline editing)"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        task_type = get_object_or_404(TaskType, task_type_id=type_id)
        data = json.loads(request.body)
        
        old_values = {
            'name': task_type.name,
            'points': task_type.points,
            'group': task_type.group.group_id,
            'is_active': task_type.is_active
        }
        
        # Update fields
        if 'name' in data:
            task_type.name = data['name'].strip()
        if 'description' in data:
            task_type.description = data['description'].strip()
        if 'group' in data:
            group = get_object_or_404(TaskGroup, group_id=data['group'])
            task_type.group = group
        if 'points' in data:
            task_type.points = int(data['points'])
        if 'price' in data:
            task_type.price = data['price'] if data['price'] else None
        if 'is_active' in data:
            task_type.is_active = data['is_active']
        
        # Advanced fields
        if 'requires_evidence' in data:
            task_type.requires_evidence = data['requires_evidence']
        if 'requires_approval' in data:
            task_type.requires_approval = data['requires_approval']
        if 'auto_complete' in data:
            task_type.auto_complete = data['auto_complete']
        
        task_type.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='update',
            model_type='tasktype',
            object_id=task_type.task_type_id,
            object_repr=f"TaskType: {task_type.name}",
            changes={'old': old_values, 'new': data},
            request=request
        )
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_delete_task_type(request, type_id):
    """AJAX: Delete task type (check usage first)"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        task_type = get_object_or_404(TaskType, task_type_id=type_id)
        
        # Check if used in any tasks
        usage_count = task_type.tasks.count()
        
        if usage_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'Cannot delete. This task type is used in {usage_count} task(s).',
                'usage_count': usage_count
            }, status=400)
        
        # Log action before delete
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='delete',
            model_type='tasktype',
            object_id=task_type.task_type_id,
            object_repr=f"TaskType: {task_type.name}",
            request=request
        )
        
        task_type.delete()
        
        return JsonResponse({'success': True, 'message': 'Task type deleted successfully.'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_reorder_task_type(request, type_id):
    """AJAX: Move task type up or down"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        task_type = get_object_or_404(TaskType, task_type_id=type_id)
        data = json.loads(request.body)
        direction = data.get('direction')
        
        if direction == 'up':
            previous = TaskType.objects.filter(
                group=task_type.group,
                order__lt=task_type.order
            ).order_by('-order').first()
            
            if previous:
                task_type.order, previous.order = previous.order, task_type.order
                task_type.save()
                previous.save()
        
        elif direction == 'down':
            next_type = TaskType.objects.filter(
                group=task_type.group,
                order__gt=task_type.order
            ).order_by('order').first()
            
            if next_type:
                task_type.order, next_type.order = next_type.order, task_type.order
                task_type.save()
                next_type.save()
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    



# ============================================
# MANAGE STAFF
# ============================================

@login_required
def admin_manage_staff_page(request):
    """Admin manages staff with search/filter"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin only.')
        return redirect('dashboard:admin_dashboard')
    
    # Get filters
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', 'all')
    branch_filter = request.GET.get('branch', 'all')
    status_filter = request.GET.get('status', 'all')
    
    # Base queryset
    staff = User.objects.all().order_by('-date_joined')
    
    # Apply search
    if search_query:
        staff = staff.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Apply filters
    if role_filter != 'all':
        staff = staff.filter(role=role_filter)
    
    if branch_filter != 'all':
        staff = staff.filter(branch=branch_filter)
    
    if status_filter == 'active':
        staff = staff.filter(is_active=True)
    elif status_filter == 'inactive':
        staff = staff.filter(is_active=False)
    
    # Get statistics
    stats = {
        'total': User.objects.count(),
        'active': User.objects.filter(is_active=True).count(),
        'inactive': User.objects.filter(is_active=False).count(),
        'staff_role': User.objects.filter(role='staff').count(),
        'managers': User.objects.filter(role='manager').count(),
        'admins': User.objects.filter(role='admin').count(),
    }
    
    # Pagination (50 per page)
    from django.core.paginator import Paginator
    paginator = Paginator(staff, 50)
    page_number = request.GET.get('page')
    staff_page = paginator.get_page(page_number)
    
    context = {
        'staff_list': staff_page,
        'stats': stats,
        'search_query': search_query,
        'role_filter': role_filter,
        'branch_filter': branch_filter,
        'status_filter': status_filter,
        'branch_choices': User.BRANCH_CHOICES,
        'role_choices': User.ROLE_CHOICES,
    }
    return render(request, 'dashboard/admin/manage_staff.html', context)


@login_required
@require_http_methods(["POST"])
def ajax_create_staff(request):
    """AJAX: Create new staff member"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # Required fields
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        role = data.get('role')
        branch = data.get('branch')
        
        if not all([username, email, role, branch]):
            return JsonResponse({
                'success': False,
                'error': 'Username, email, role, and branch are required'
            }, status=400)
        
        # Check if username exists
        if User.objects.filter(username=username).exists():
            return JsonResponse({
                'success': False,
                'error': 'Username already exists'
            }, status=400)
        
        # Check if email exists
        if User.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'error': 'Email already exists'
            }, status=400)
        
        # Generate employee_id
        today = date.today()
        prefix = f"EMP{today.strftime('%y')}"
        last_emp = User.objects.filter(
            employee_id__startswith=prefix
        ).order_by('employee_id').last()
        
        if last_emp and last_emp.employee_id:
            last_num = int(last_emp.employee_id[5:])
            new_num = last_num + 1
        else:
            new_num = 1
        
        employee_id = f"{prefix}{new_num:04d}"
        
        # Generate random password
        password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
        
        # Create user
        user = User.objects.create(
            username=username,
            email=email,
            employee_id=employee_id,
            role=role,
            branch=branch,
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
            phone=data.get('phone', ''),
            password=make_password(password),
            is_active=True,
            is_staff=True if role == 'admin' else False,
        )
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='create',
            model_type='user',
            object_id=employee_id,
            object_repr=f"User: {username}",
            changes={
                'username': username,
                'email': email,
                'role': role,
                'branch': branch
            },
            notes=f"Generated employee_id: {employee_id}",
            request=request
        )
        
        # Send email with credentials (optional)
        try:
            send_mail(
                subject='CatzoTeam Account Created',
                message=f'''
Hello {user.first_name or username},

Your CatzoTeam account has been created!

Username: {username}
Password: {password}
Employee ID: {employee_id}

Please change your password after first login.

Best regards,
CatzoTeam Admin
''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Staff created successfully!',
            'employee_id': employee_id,
            'password': password,  # Show to admin once
            'user': {
                'id': user.id,
                'username': username,
                'email': email,
                'employee_id': employee_id,
                'role': role,
                'branch': branch
            }
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_update_staff(request, user_id):
    """AJAX: Update staff member"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        user = get_object_or_404(User, id=user_id)
        data = json.loads(request.body)
        
        old_values = {
            'role': user.role,
            'branch': user.branch,
            'is_active': user.is_active,
            'email': user.email
        }
        
        # Update fields
        if 'first_name' in data:
            user.first_name = data['first_name'].strip()
        if 'last_name' in data:
            user.last_name = data['last_name'].strip()
        if 'email' in data:
            email = data['email'].strip()
            # Check if email exists for other users
            if User.objects.filter(email=email).exclude(id=user_id).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Email already in use'
                }, status=400)
            user.email = email
        
        if 'phone' in data:
            user.phone = data['phone'].strip()
        if 'role' in data:
            user.role = data['role']
            user.is_staff = True if data['role'] == 'admin' else False
        if 'branch' in data:
            user.branch = data['branch']
        if 'is_active' in data:
            user.is_active = data['is_active']
        
        user.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='update',
            model_type='user',
            object_id=user.employee_id,
            object_repr=f"User: {user.username}",
            changes={'old': old_values, 'new': data},
            request=request
        )
        
        return JsonResponse({'success': True, 'message': 'Staff updated successfully'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_soft_delete_staff(request, user_id):
    """AJAX: Soft delete (deactivate) staff"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Prevent self-deletion
        if user.id == request.user.id:
            return JsonResponse({
                'success': False,
                'error': 'Cannot deactivate your own account'
            }, status=400)
        
        user.is_active = False
        user.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='deactivate',
            model_type='user',
            object_id=user.employee_id,
            object_repr=f"User: {user.username}",
            request=request
        )
        
        return JsonResponse({'success': True, 'message': f'{user.username} deactivated'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_activate_staff(request, user_id):
    """AJAX: Reactivate staff"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        user.is_active = True
        user.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='activate',
            model_type='user',
            object_id=user.employee_id,
            object_repr=f"User: {user.username}",
            request=request
        )
        
        return JsonResponse({'success': True, 'message': f'{user.username} activated'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_reset_staff_password(request, user_id):
    """AJAX: Reset staff password"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Generate new random password
        new_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
        
        user.password = make_password(new_password)
        user.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='password_reset',
            model_type='user',
            object_id=user.employee_id,
            object_repr=f"User: {user.username}",
            request=request
        )
        
        # Send email
        try:
            send_mail(
                subject='CatzoTeam Password Reset',
                message=f'''
Hello {user.first_name or user.username},

Your password has been reset by an administrator.

New Password: {new_password}

Please change your password after login.

Best regards,
CatzoTeam Admin
''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                fail_silently=True,
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Password reset successfully',
            'new_password': new_password  # Show to admin
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ajax_staff_performance(request, user_id):
    """AJAX: Get staff performance/points history"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Get performance data
        from performance.models import DailyPoints, MonthlyIncentive
        from task_management.models import Task
        
        # Last 30 days points
        thirty_days_ago = date.today() - timedelta(days=30)
        daily_points = DailyPoints.objects.filter(
            user=user,
            date__gte=thirty_days_ago
        ).order_by('-date')[:30]
        
        # Current month
        first_day = date.today().replace(day=1)
        monthly = MonthlyIncentive.objects.filter(
            user=user,
            month=first_day
        ).first()
        
        # Task statistics
        task_stats = {
            'total': Task.objects.filter(assigned_staff=user).count(),
            'completed': Task.objects.filter(assigned_staff=user, status='completed').count(),
            'in_progress': Task.objects.filter(assigned_staff=user, status='in_progress').count(),
            'pending': Task.objects.filter(assigned_staff=user, status='assigned').count(),
        }
        
        # Points breakdown
        points_data = {
            'dates': [dp.date.strftime('%Y-%m-%d') for dp in daily_points],
            'points': [float(dp.points) for dp in daily_points],
        }
        
        return JsonResponse({
            'success': True,
            'performance': {
                'task_stats': task_stats,
                'monthly_points': float(monthly.total_points) if monthly else 0,
                'daily_points': points_data,
                'total_incentive': float(monthly.total_incentive) if monthly and hasattr(monthly, 'total_incentive') else 0,
            }
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_reassign_tasks(request, user_id):
    """AJAX: Reassign all tasks from one staff to another"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        from_user = get_object_or_404(User, id=user_id)
        data = json.loads(request.body)
        to_user_id = data.get('to_user_id')
        
        if not to_user_id:
            return JsonResponse({
                'success': False,
                'error': 'Target staff member required'
            }, status=400)
        
        to_user = get_object_or_404(User, id=to_user_id)
        
        # Get all non-completed tasks
        tasks_to_reassign = Task.objects.filter(
            assigned_staff=from_user,
            status__in=['assigned', 'in_progress']
        )
        
        count = tasks_to_reassign.count()
        tasks_to_reassign.update(assigned_staff=to_user)
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='task_reassign',
            model_type='task',
            object_id=f"{from_user.employee_id}_to_{to_user.employee_id}",
            object_repr=f"Reassigned {count} tasks",
            changes={
                'from': from_user.username,
                'to': to_user.username,
                'count': count
            },
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{count} tasks reassigned from {from_user.username} to {to_user.username}'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
# ============================================
# MANAGE TASKS - TASK TYPES & GROUPS
# ============================================

@login_required
def admin_manage_tasks_page(request):
    """Admin manages task types and groups"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin only.')
        return redirect('dashboard:admin_dashboard')
    
    # Get all groups with task type counts
    groups = TaskGroup.objects.annotate(
        task_count=Count('task_types')
    ).order_by('order', 'name')
    
    # Get all task types with usage count
    task_types = TaskType.objects.annotate(
        usage_count=Count('tasks')
    ).select_related('group').order_by('group__order', 'order', 'name')
    
    # Get selected group from query param
    selected_group = request.GET.get('group', 'all')
    
    # Get the selected group name
    selected_group_name = ""
    if selected_group != 'all':
        try:
            selected_group_obj = TaskGroup.objects.get(group_id=selected_group)  # Changed here
            selected_group_name = selected_group_obj.name
        except TaskGroup.DoesNotExist:
            selected_group_name = ""
    
    # Filter types by group if selected
    if selected_group != 'all':
        task_types = task_types.filter(group__group_id=selected_group)  # Also changed here
    
    # Check if "Uncategorized" group exists, create if not
    uncategorized_group, created = TaskGroup.objects.get_or_create(
        name='Uncategorized',
        defaults={
            'description': 'Uncategorized tasks',
            'is_active': True,
            'order': 9999
        }
    )
    
    context = {
        'groups': groups,
        'task_types': task_types,
        'selected_group': selected_group,
        'uncategorized_group': uncategorized_group,
        'selected_group_name': selected_group_name,
    }
    return render(request, 'task_management/admin/manage_tasks.html', context)

@login_required
@require_http_methods(["POST"])
def ajax_create_task_group(request):
    """AJAX: Create new task group"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)
        
        # Check if name exists
        if TaskGroup.objects.filter(name=name).exists():
            return JsonResponse({'success': False, 'error': 'Group name already exists'}, status=400)
        
        # Get max order + 1
        max_order = TaskGroup.objects.aggregate(models.Max('order'))['order__max'] or 0
        
        group = TaskGroup.objects.create(
            name=name,
            description=description,
            is_active=True,
            order=max_order + 1
        )
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='create',
            model_type='taskgroup',
            object_id=group.group_id,
            object_repr=f"TaskGroup: {group.name}",
            changes={'name': name, 'description': description},
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'group': {
                'id': group.group_id,
                'name': group.name,
                'description': group.description,
                'task_count': 0
            }
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_update_task_group(request, group_id):
    """AJAX: Update task group"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        group = get_object_or_404(TaskGroup, group_id=group_id)
        data = json.loads(request.body)
        
        old_values = {
            'name': group.name,
            'description': group.description,
            'is_active': group.is_active
        }
        
        # Update fields
        if 'name' in data:
            name = data['name'].strip()
            if not name:
                return JsonResponse({'success': False, 'error': 'Name cannot be empty'}, status=400)
            group.name = name
        
        if 'description' in data:
            group.description = data['description'].strip()
        
        if 'is_active' in data:
            group.is_active = data['is_active']
        
        group.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='update',
            model_type='taskgroup',
            object_id=group.group_id,
            object_repr=f"TaskGroup: {group.name}",
            changes={'old': old_values, 'new': {
                'name': group.name,
                'description': group.description,
                'is_active': group.is_active
            }},
            request=request
        )
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_delete_task_group(request, group_id):
    """AJAX: Delete task group (moves types to Uncategorized)"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        group = get_object_or_404(TaskGroup, group_id=group_id)
        
        # Get or create Uncategorized group
        uncategorized_group, created = TaskGroup.objects.get_or_create(
            name='Uncategorized',
            defaults={
                'description': 'Uncategorized tasks',
                'is_active': True,
                'order': 9999
            }
        )
        
        # Move all task types to Uncategorized
        moved_count = group.task_types.update(group=uncategorized_group)
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='delete',
            model_type='taskgroup',
            object_id=group.group_id,
            object_repr=f"TaskGroup: {group.name}",
            notes=f"Moved {moved_count} task types to Uncategorized",
            request=request
        )
        
        group.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Group deleted. {moved_count} task types moved to Uncategorized.'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_reorder_group(request, group_id):
    """AJAX: Move group up or down"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        group = get_object_or_404(TaskGroup, group_id=group_id)
        data = json.loads(request.body)
        direction = data.get('direction')  # 'up' or 'down'
        
        if direction == 'up':
            # Swap with previous
            previous = TaskGroup.objects.filter(order__lt=group.order).order_by('-order').first()
            if previous:
                group.order, previous.order = previous.order, group.order
                group.save()
                previous.save()
        
        elif direction == 'down':
            # Swap with next
            next_group = TaskGroup.objects.filter(order__gt=group.order).order_by('order').first()
            if next_group:
                group.order, next_group.order = next_group.order, group.order
                group.save()
                next_group.save()
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# TASK TYPE AJAX OPERATIONS
# ============================================

@login_required
@require_http_methods(["POST"])
def ajax_create_task_type(request):
    """AJAX: Create new task type"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # Required fields
        name = data.get('name', '').strip()
        group_id = data.get('group')
        points = data.get('points', 0)
        
        if not name or not group_id:
            return JsonResponse({'success': False, 'error': 'Name and group are required'}, status=400)
        
        group = get_object_or_404(TaskGroup, group_id=group_id)
        
        # Get max order in this group
        max_order = TaskType.objects.filter(group=group).aggregate(
            Max('order')
        )['order__max'] or 0
        
        task_type = TaskType.objects.create(
            name=name,
            group=group,
            points=int(points),
            description=data.get('description', ''),
            price=data.get('price') if data.get('price') else None,
            category=data.get('category', ''),
            is_active=data.get('is_active', True),
            order=max_order + 1,
            # Advanced fields
            rule_type=data.get('rule_type', ''),
            price_min=data.get('price_min'),
            price_max=data.get('price_max'),
            count_min=data.get('count_min'),
            count_max=data.get('count_max'),
            requires_evidence=data.get('requires_evidence', False),
            requires_approval=data.get('requires_approval', False),
            auto_complete=data.get('auto_complete', True),
        )
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='create',
            model_type='tasktype',
            object_id=task_type.task_type_id,
            object_repr=f"TaskType: {task_type.name}",
            changes=data,
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'task_type': {
                'id': task_type.task_type_id,
                'name': task_type.name,
                'points': task_type.points
            }
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_update_task_type(request, type_id):
    """AJAX: Update task type (inline editing)"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        task_type = get_object_or_404(TaskType, task_type_id=type_id)
        data = json.loads(request.body)
        
        old_values = {
            'name': task_type.name,
            'points': task_type.points,
            'group': task_type.group.group_id,
            'is_active': task_type.is_active
        }
        
        # Update fields
        if 'name' in data:
            task_type.name = data['name'].strip()
        if 'description' in data:
            task_type.description = data['description'].strip()
        if 'group' in data:
            group = get_object_or_404(TaskGroup, group_id=data['group'])
            task_type.group = group
        if 'points' in data:
            task_type.points = int(data['points'])
        if 'price' in data:
            task_type.price = data['price'] if data['price'] else None
        if 'is_active' in data:
            task_type.is_active = data['is_active']
        
        # Advanced fields
        if 'requires_evidence' in data:
            task_type.requires_evidence = data['requires_evidence']
        if 'requires_approval' in data:
            task_type.requires_approval = data['requires_approval']
        if 'auto_complete' in data:
            task_type.auto_complete = data['auto_complete']
        
        task_type.save()
        
        # Log action
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='update',
            model_type='tasktype',
            object_id=task_type.task_type_id,
            object_repr=f"TaskType: {task_type.name}",
            changes={'old': old_values, 'new': data},
            request=request
        )
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_delete_task_type(request, type_id):
    """AJAX: Delete task type (check usage first)"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        task_type = get_object_or_404(TaskType, task_type_id=type_id)
        
        # Check if used in any tasks
        usage_count = task_type.tasks.count()
        
        if usage_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'Cannot delete. This task type is used in {usage_count} task(s).',
                'usage_count': usage_count
            }, status=400)
        
        # Log action before delete
        from task_management.models import log_admin_action
        log_admin_action(
            user=request.user,
            action='delete',
            model_type='tasktype',
            object_id=task_type.task_type_id,
            object_repr=f"TaskType: {task_type.name}",
            request=request
        )
        
        task_type.delete()
        
        return JsonResponse({'success': True, 'message': 'Task type deleted successfully.'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ajax_reorder_task_type(request, type_id):
    """AJAX: Move task type up or down"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        task_type = get_object_or_404(TaskType, task_type_id=type_id)
        data = json.loads(request.body)
        direction = data.get('direction')
        
        if direction == 'up':
            previous = TaskType.objects.filter(
                group=task_type.group,
                order__lt=task_type.order
            ).order_by('-order').first()
            
            if previous:
                task_type.order, previous.order = previous.order, task_type.order
                task_type.save()
                previous.save()
        
        elif direction == 'down':
            next_type = TaskType.objects.filter(
                group=task_type.group,
                order__gt=task_type.order
            ).order_by('order').first()
            
            if next_type:
                task_type.order, next_type.order = next_type.order, task_type.order
                task_type.save()
                next_type.save()
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)